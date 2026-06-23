#!/usr/bin/env python3
"""
生成诗词卡片 Excel 打印模板 — 6 卡版
布局：A4 横向，底部 4 张竖卡 (7.4×12.6cm)，顶部 2 张横卡 (旋转 90°)

页面尺寸精确匹配 A4 横向 (297×210mm)，底部 4 卡无页边距、直接裁切。
顶部 2 卡居中于剩余空间内，裁切后与底部卡片等大。
"""
import json
import argparse
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


def split_long_lines(content_str, max_len=9):
    """按中文标点拆分超长行。"""
    if not content_str:
        return []
    raw_lines = content_str.split('\n')
    processed = []
    for line in raw_lines:
        line = line.strip()
        if not line:
            continue
        if len(line) <= max_len:
            processed.append(line)
        else:
            parts = re.split(r'(，|；|、|。)', line)
            current = ""
            for i in range(0, len(parts), 2):
                chunk = parts[i]
                if i + 1 < len(parts):
                    chunk += parts[i+1]
                if len(current) + len(chunk) <= max_len and current != "":
                    current += chunk
                else:
                    if current:
                        processed.append(current)
                    current = chunk
            if current:
                processed.append(current)
    return processed


# ═══════════════════════════════════════════════════════════════════
#  布局常量 — 精确匹配 A4 横向 297×210mm
# ═══════════════════════════════════════════════════════════════════
#
#  内容尺寸故意设为 A4 的 ~105%（宽高比与 A4 完全一致 = 1.414:1），
#  由 fitToPage 等比缩放至纸张，确保内容铺满整个可打印区域，
#  消除因内容偏小导致的空白边距。
#
#  ┌─────────────────┬─────────────────┐  行 1-2   (裁切余量)
#  │   顶部横卡 #5    │   顶部横卡 #6    │  行 3-20  (18 行 ≈ 74mm)
#  │   (cols 1-2)     │   (cols 3-4)     │
#  ├────────┬────────┼────────┬────────┤  行 21
#  │ 竖卡#1 │ 竖卡#2 │ 竖卡#3 │ 竖卡#4 │  行 21-50 (30 行 = 126mm)
#  │ 7.4cm  │ 7.4cm  │ 7.4cm  │ 7.4cm  │  ← 四等分长边，直接裁切
#  └────────┴────────┴────────┴────────┘  行 50 = 纸张底边

CARD_COLS = 6
CARDS_PER_PAGE = 6
BOTTOM_CARDS = 4
TOP_CARDS = 2

# 对称列宽（厘米）：[2.25, 5.175, 7.425, 7.425, 5.175, 2.25]
# 转换为 Excel 字符宽度以达到 96 DPI 精确像素：
# 1122.5 像素分配为 [85.04, 195.59, 280.625, 280.625, 195.59, 85.04]
COL_WIDTHS = [11.434, 27.227, 39.375, 39.375, 27.227, 11.434]

BOTTOM_ROW_SPAN = 30          # 底部卡片：30 行
TOP_ROW_SPAN = 20             # 顶部区域：20 行
TOP_CARD_MARGIN = 1           # 顶部裁切余量 (行 1 作为边距)
TOP_CARD_ROW_SPAN = 18        # 顶部卡片实际高度 (行 2 至 19)
TOTAL_ROWS = BOTTOM_ROW_SPAN + TOP_ROW_SPAN  # 50

COLOR_TITLE = "1A1A1A"
COLOR_AUTHOR = "595959"
COLOR_DECO = "C0392B"
COLOR_PAGE = "8C8C8C"
COLOR_CONTENT = "1A1A1A"

# 底部竖卡正面的字号/行距阶梯
LAYOUTS = [
    # (font_size, rows_per_line, max_len)
    (22, 6, 9),
    (22, 5, 9),
    (20, 4, 10),
    (18, 3, 12),
    (16, 2, 14),
    (14, 2, 16),
    (14, 1, 16),
    (12, 1, 18),
    (10, 1, 22),
    (9, 1, 25),
    (8, 1, 28)
]


def find_best_layout(content_str):
    """为底部竖卡选择最佳字号与行距。"""
    available = BOTTOM_ROW_SPAN - 3
    for font_size, rows_per_line, max_len in LAYOUTS:
        lines = split_long_lines(content_str, max_len)
        total_rows = len(lines) * rows_per_line
        if total_rows <= available:
            return font_size, rows_per_line, max_len, lines
    lines = split_long_lines(content_str, 28)
    return 8, 1, 28, lines


def get_top_font_size(content_str):
    """为顶部横卡（旋转 90° 单元格）选择合适字号。"""
    if not content_str:
        return 18
    chars = len(content_str.replace('\n', '').replace(' ', ''))
    if chars <= 30:
        return 20
    elif chars <= 50:
        return 18
    elif chars <= 80:
        return 16
    elif chars <= 120:
        return 14
    elif chars <= 180:
        return 12
    else:
        return 10


# ═══════════════════════════════════════════════════════════════════
#  Sheet 初始化
# ═══════════════════════════════════════════════════════════════════

def setup_sheet(ws):
    """配置 A4 横向、零页边距、行高列宽精确铺满整页。"""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 必须零页边距，且把 header 和 footer 也设为 0，否则 Excel 会强行留出边距
    ws.page_margins = PageMargins(left=0, right=0, top=0, bottom=0, header=0, footer=0)
    ws.print_options.horizontalCentered = True

    # 设置列宽
    for c, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # 精确行高设置：总高 595.27 磅 (21.0cm)
    ws.row_dimensions[1].height = 13.82  # 顶部外边距 (第 1 行)
    for r in range(2, 20):
        ws.row_dimensions[r].height = 11.693  # 顶部横卡 (18行 ≈ 7.425cm)
    ws.row_dimensions[20].height = 13.82  # 顶部与底部之间的间隙/余量 (第 20 行)
    for r in range(21, 51):
        ws.row_dimensions[r].height = 11.905  # 底部竖卡 (30行 ≈ 12.6cm)


# ═══════════════════════════════════════════════════════════════════
#  卡片定位
# ═══════════════════════════════════════════════════════════════════

def get_card_position_front(index):
    """返回 (col_start, row_start, col_span, row_span, is_top)。"""
    if index < BOTTOM_CARDS:
        # 底部 4 张竖卡，利用对称 6 列合并出等宽 of 4 张卡片：
        # 卡 1: Col 1-2 (2.25 + 5.175 = 7.425 cm)
        # 卡 2: Col 3 (7.425 cm)
        # 卡 3: Col 4 (7.425 cm)
        # 卡 4: Col 5-6 (5.175 + 2.25 = 7.425 cm)
        if index == 0:
            col_start, col_span = 1, 2
        elif index == 1:
            col_start, col_span = 3, 1
        elif index == 2:
            col_start, col_span = 4, 1
        else:
            col_start, col_span = 5, 2
        return (col_start,
                TOP_ROW_SPAN + 1,       # 底部起始行 = 21
                col_span,
                BOTTOM_ROW_SPAN,         # 30 行
                False)
    else:
        # 顶部 2 张横卡 (旋转90°)：
        # 卡 5: Col 2-3 (5.175 + 7.425 = 12.6 cm)
        # 卡 6: Col 4-5 (7.425 + 5.175 = 12.6 cm)
        # 左右留空 Col 1 (2.25 cm) 和 Col 6 (2.25 cm) 作裁切余量
        top_idx = index - BOTTOM_CARDS
        if top_idx == 0:
            col_start = 2
        else:
            col_start = 4
        return (col_start,
                TOP_CARD_MARGIN + 1,     # 顶部卡片起始行 = 2
                2,                       # 跨 2 列
                TOP_CARD_ROW_SPAN,       # 18 行
                True)


def get_card_position_back(index):
    """返回背面（双面打印沿短边翻转）的镜像位置。"""
    col, row, span, rspan, is_top = get_card_position_front(index)
    # 对于 6 列，镜像列公式：mirrored_col_start = CARD_COLS - col - span + 2
    mirrored_col = CARD_COLS - col - span + 2
    return mirrored_col, row, span, rspan, is_top


# ═══════════════════════════════════════════════════════════════════
#  裁切边框
# ═══════════════════════════════════════════════════════════════════

def add_card_border(ws, col_start, row_start, row_span, col_span=1,
                    color="8C8C8C", style="dashed"):
    """为卡片区域绘制虚线裁切框。"""
    s = Side(style=style, color=color)
    row_end = row_start + row_span - 1
    col_end = col_start + col_span - 1

    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            t = s if r == row_start else None
            b = s if r == row_end else None
            le = s if c == col_start else None
            ri = s if c == col_end else None
            if t or b or le or ri:
                ws.cell(row=r, column=c).border = Border(
                    top=t, bottom=b, left=le, right=ri)


# ═══════════════════════════════════════════════════════════════════
#  正面内容（诗文）
# ═══════════════════════════════════════════════════════════════════

def fill_card_front_bottom(ws, col_start, row_start, col_span, row_span, poem):
    """底部竖卡正面：序号 + 居中诗文。"""
    add_card_border(ws, col_start, row_start, row_span, col_span=col_span)
    if not poem.get('title'):
        return

    # 序号（顶部 3 行合并，横跨 col_span）
    ws.merge_cells(start_row=row_start, start_column=col_start,
                   end_row=row_start + 2, end_column=col_start + col_span - 1)
    cell = ws.cell(row=row_start, column=col_start)
    cell.value = f"·  {poem['seq']}  ·"
    cell.font = Font(name='宋体', size=12, color=COLOR_DECO, italic=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # 诗文内容
    font_size, rows_per_line, _, lines = find_best_layout(poem.get('content', ''))
    if not lines:
        return

    available = row_span - 3
    total = len(lines) * rows_per_line
    offset = (available - total) // 2
    cur = row_start + 3 + offset

    for line in lines:
        cell = ws.cell(row=cur, column=col_start)
        if rows_per_line > 1 or col_span > 1:
            ws.merge_cells(start_row=cur, start_column=col_start,
                           end_row=cur + rows_per_line - 1, end_column=col_start + col_span - 1)
        cell.value = line
        cell.font = Font(name='宋体', size=font_size, color=COLOR_CONTENT, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cur += rows_per_line


def fill_card_front_top(ws, col_start, row_start, col_span, row_span, poem):
    """顶部横卡正面：合并 col_span 列，文字旋转 90°。"""
    add_card_border(ws, col_start, row_start, row_span, col_span=col_span)
    if not poem.get('title'):
        return

    ws.merge_cells(start_row=row_start, start_column=col_start,
                   end_row=row_start + row_span - 1, end_column=col_start + col_span - 1)

    content = poem.get('content', '')
    font_size = get_top_font_size(content)
    max_len = 10 if font_size >= 16 else 16
    lines = split_long_lines(content, max_len)

    text = f"·  {poem['seq']}  ·\n\n" + "\n".join(lines)

    cell = ws.cell(row=row_start, column=col_start)
    cell.value = text
    cell.font = Font(name='宋体', size=font_size, color=COLOR_CONTENT, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                               text_rotation=90, wrap_text=True)


# ═══════════════════════════════════════════════════════════════════
#  背面内容（题目 / 作者 / 页码）
# ═══════════════════════════════════════════════════════════════════

def fill_card_back_bottom(ws, col_start, row_start, col_span, row_span, poem):
    """底部竖卡背面：序号、装饰线、标题、作者、页码。"""
    add_card_border(ws, col_start, row_start, row_span, col_span=col_span)
    if not poem.get('title'):
        return

    # NO. xxx
    r = row_start + 3
    if col_span > 1:
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_start + col_span - 1)
    c = ws.cell(row=r, column=col_start)
    c.value = f"NO. {poem['seq']:03d}"
    c.font = Font(name='Arial', size=12, color=COLOR_DECO, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # —— ✦ ——
    r = row_start + 5
    if col_span > 1:
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_start + col_span - 1)
    c = ws.cell(row=r, column=col_start)
    c.value = "—— ✦ ——"
    c.font = Font(name='宋体', size=14, color=COLOR_DECO)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # 标题（大字，7 行合并）
    r = row_start + 9
    ws.merge_cells(start_row=r, start_column=col_start,
                   end_row=r + 6, end_column=col_start + col_span - 1)
    c = ws.cell(row=r, column=col_start)
    c.value = poem["title"]
    c.font = Font(name='宋体', size=24, color=COLOR_TITLE, bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 作者
    r = row_start + 18
    ws.merge_cells(start_row=r, start_column=col_start,
                   end_row=r + 1, end_column=col_start + col_span - 1)
    c = ws.cell(row=r, column=col_start)
    c.value = poem["author"]
    c.font = Font(name='宋体', size=16, color=COLOR_AUTHOR)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # — —
    r = row_start + 22
    if col_span > 1:
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_start + col_span - 1)
    c = ws.cell(row=r, column=col_start)
    c.value = "— —"
    c.font = Font(name='宋体', size=12, color=COLOR_DECO)
    c.alignment = Alignment(horizontal='center', vertical='center')

    # 《课本》页码
    r = row_start + 25
    if col_span > 1:
        ws.merge_cells(start_row=r, start_column=col_start, end_row=r, end_column=col_start + col_span - 1)
    c = ws.cell(row=r, column=col_start)
    c.value = f"《课本》{poem['page']}"
    c.font = Font(name='宋体', size=12, color=COLOR_PAGE)
    c.alignment = Alignment(horizontal='center', vertical='center')


def fill_card_back_top(ws, col_start, row_start, col_span, row_span, poem):
    """顶部横卡背面：合并 col_span 列，文字旋转 90°。"""
    add_card_border(ws, col_start, row_start, row_span, col_span=col_span)
    if not poem.get('title'):
        return

    ws.merge_cells(start_row=row_start, start_column=col_start,
                   end_row=row_start + row_span - 1, end_column=col_start + col_span - 1)

    text = (f"NO. {poem['seq']:03d}\n"
            f"—— ✦ ——\n\n"
            f"{poem['title']}\n\n"
            f"{poem['author']}\n"
            f"— —\n"
            f"《课本》{poem['page']}")

    cell = ws.cell(row=row_start, column=col_start)
    cell.value = text
    cell.font = Font(name='宋体', size=18, color=COLOR_TITLE, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center',
                               text_rotation=90, wrap_text=True)


# ═══════════════════════════════════════════════════════════════════
#  主入口
# ═══════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="生成诗词卡片 Excel 模板（6 卡/页）")
    parser.add_argument("start", type=int, help="起始序号")
    parser.add_argument("end", type=int, help="结束序号")
    args = parser.parse_args()

    if args.start < 1 or args.end < args.start:
        print("错误：范围无效")
        return

    count = args.end - args.start + 1
    if count % CARDS_PER_PAGE != 0:
        print(f"错误：生成的卡片数量必须是 {CARDS_PER_PAGE} 的倍数，"
              f"当前范围包含 {count} 首。")
        return

    try:
        with open("poems_index.json", "r", encoding="utf-8") as f:
            all_poems = json.load(f)
    except Exception as e:
        print(f"读取 poems_index.json 失败: {e}")
        return

    poems = [p for p in all_poems if args.start <= p["seq"] <= args.end]

    if len(poems) < count:
        print(f"警告：JSON 中找到 {len(poems)} 首，少于请求的 {count} 首，用空白补齐。")

    empty = {"seq": 0, "title": "", "author": "", "page": "", "content": ""}
    while len(poems) % CARDS_PER_PAGE != 0 or len(poems) < count:
        poems.append(empty)

    wb = Workbook()
    wb.remove(wb.active)

    chunks = [poems[i:i + CARDS_PER_PAGE]
              for i in range(0, len(poems), CARDS_PER_PAGE)]

    for chunk in chunks:
        valid = [p['seq'] for p in chunk if p['seq'] != 0]
        if not valid:
            continue
        label = f"{valid[0]}-{valid[-1]}"

        # ── 正面 ──
        ws_f = wb.create_sheet(f"正面({label})")
        setup_sheet(ws_f)
        for i, poem in enumerate(chunk):
            cs, rs, span, rspan, is_top = get_card_position_front(i)
            if is_top:
                fill_card_front_top(ws_f, cs, rs, span, rspan, poem)
            else:
                fill_card_front_bottom(ws_f, cs, rs, span, rspan, poem)

        # ── 背面 ──
        ws_b = wb.create_sheet(f"背面({label})")
        setup_sheet(ws_b)
        for i, poem in enumerate(chunk):
            cs, rs, span, rspan, is_top = get_card_position_back(i)
            if is_top:
                fill_card_back_top(ws_b, cs, rs, span, rspan, poem)
            else:
                fill_card_back_bottom(ws_b, cs, rs, span, rspan, poem)

    out_file = f"poem_cards_{args.start}-{args.end}.xlsx"
    wb.save(out_file)
    print(f"✓ 成功生成: {out_file} (共 {len(chunks)} 页双面A4)")


if __name__ == "__main__":
    main()